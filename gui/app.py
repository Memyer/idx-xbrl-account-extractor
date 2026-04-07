"""
gui/app.py — IDX Superapp (CustomTkinter)
Redesigned UI/UX — lebih user-friendly, clean, dan intuitif.
"""
import customtkinter as ctk
from tkinter import filedialog, messagebox
import tkinter as tk
import threading
from datetime import datetime
import os, sys, shutil

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from core.link_generator import generate_links, PERIOD_MAP
from core.downloader import download_all
from core.unzipper import unzip_all
from core.xbrl_extractor import extract_all, ALL_METRICS, DEFAULT_METRICS
from core.data_cleaner import clean_data
from core.amount_normalizer import normalize_to_full_idr, ROUNDING_COLUMN, CURRENCY_COLUMN
from core.csv_exporter import export_to_excel
from core.metric_calculator import calculate_metrics

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# ── Palette — Modern Obsidian ─────────────────────────────────────────
# Inspirasi: Linear, Vercel, Notion dark mode
C_BG        = "#09090b"   # zinc-950 — background utama, hampir hitam
C_SIDEBAR   = "#0f0f12"   # sidebar sedikit lebih terang dari bg
C_SURFACE   = "#18181b"   # zinc-900 — card/panel surface
C_CARD      = "#1f1f23"   # zinc-800/900 blend — card dalam surface
C_CARD2     = "#141417"   # lebih gelap untuk nested element
C_BORDER    = "#27272a"   # zinc-800 — border halus
C_BORDER2   = "#3f3f46"   # zinc-700 — border lebih visible

# Accent — Electric Indigo (lebih premium dari biru biasa)
C_ACCENT    = "#6366f1"   # indigo-500
C_ACCENT2   = "#4f46e5"   # indigo-600 (hover)
C_ACCENT_LT = "#818cf8"   # indigo-400 (text on dark)

# Semantic
C_SUCCESS   = "#22c55e"   # green-500
C_SUCCESS2  = "#16a34a"   # green-600
C_WARNING   = "#f59e0b"   # amber-500
C_DANGER    = "#ef4444"   # red-500
C_DANGER2   = "#dc2626"   # red-600

# Typography
C_TEXT      = "#fafafa"   # zinc-50 — teks utama, hampir putih
C_TEXT2     = "#d4d4d8"   # zinc-300 — teks sekunder
C_MUTED     = "#71717a"   # zinc-500 — placeholder/hint
C_MUTED2    = "#a1a1aa"   # zinc-400 — label sekunder
C_WHITE     = "#ffffff"

# Accent warna untuk preset
C_PURPLE    = "#a78bfa"   # violet-400
C_PURPLE2   = "#7c3aed"   # violet-700 (bg)
C_GREEN     = "#34d399"   # emerald-400
C_GREEN2    = "#065f46"   # emerald-900 (bg)

# Step states
C_STEP_WAIT = "#27272a"
C_STEP_RUN  = "#f59e0b"
C_STEP_OK   = "#22c55e"
C_STEP_ERR  = "#ef4444"

NORMALIZATION_REFERENCE_METRICS = [ROUNDING_COLUMN, CURRENCY_COLUMN]


# ═════════════════════════════════════════════════════════════════════
class IDXSuperApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("IDX Superapp — Ekstraksi Data Keuangan IDX")
        self.geometry("1200x800")
        self.minsize(1050, 700)
        self.configure(fg_color=C_BG)

        self._stop_flag = False
        self._running   = False

        # ── Settings vars ──
        self.var_year     = tk.IntVar(value=datetime.now().year)
        self.var_period   = tk.StringVar(value="Tahunan (Audit)")
        self.var_base_dir = tk.StringVar(value=os.getcwd())
        self.var_usd_rate = tk.StringVar(value="16000")

        self.var_companies_file = tk.StringVar()
        self.var_download_dir   = tk.StringVar()
        self.var_xbrl_dir       = tk.StringVar()
        self.var_extract_dir    = tk.StringVar()
        self._update_paths()
        self.var_base_dir.trace_add("write", lambda *_: self._update_paths())

        # ── Account selector state ──
        self._accounts_by_category, self._all_account_options = self._load_accounts_categorized()
        if not self._all_account_options:
            self._all_account_options = list(ALL_METRICS)
            self._accounts_by_category = {"Semua": self._all_account_options}
        self._active_category = "Semua"

        self._selected_metrics = [m for m in DEFAULT_METRICS if m in self._all_account_options]
        if not self._selected_metrics and self._all_account_options:
            self._selected_metrics = self._all_account_options[:min(20, len(self._all_account_options))]

        self._search_var = tk.StringVar()
        self._search_var.trace_add("write", lambda *_: self._filter_metrics())
        self._filtered_metrics = list(self._all_account_options)

        # UI refs
        self.account_listbox   = None
        self.selected_list     = None
        self.lbl_selected_count = None
        self._cat_buttons      = {}
        self._step_w           = {}
        self._sidebar_status   = {}   # key → label di sidebar

        self._build_ui()

    # ─────────────────────────────────────────────────────────────────
    @staticmethod
    def _bundle_path(filename: str) -> str:
        """Resolve path file: dari sys._MEIPASS (EXE bundle) atau folder lokal."""
        if getattr(sys, "frozen", False):
            # Saat dijalankan sebagai EXE PyInstaller
            return os.path.join(sys._MEIPASS, filename)
        # Saat dijalankan sebagai script biasa
        base = os.path.dirname(os.path.abspath(__file__))
        candidates = [
            os.path.join(base, filename),
            os.path.join(os.path.dirname(base), filename),
        ]
        return next((p for p in candidates if os.path.exists(p)), candidates[0])

    def _load_accounts_categorized(self) -> tuple:
        xlsx_path = self._bundle_path("Pos_XBRL_IDX_Lengkap.xlsx")
        if not os.path.exists(xlsx_path):
            xlsx_path = None
        SHEET_LABELS = {
            "1. Informasi Umum": "Informasi Umum",
            "2. Neraca":         "Neraca",
            "3. Laba Rugi":      "Laba Rugi",
            "4. Arus Kas":       "Arus Kas",
        }
        by_cat, all_names, seen = {}, [], set()
        if xlsx_path:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
                for sh in wb.sheetnames:
                    label = SHEET_LABELS.get(sh, sh)
                    cat_list = []
                    for row in wb[sh].iter_rows(min_row=3, values_only=True):
                        name = row[1] if row and len(row) > 1 else None
                        if not name or name == "Nama Pos": continue
                        name = str(name).strip()
                        if not name: continue
                        cat_list.append(name)
                        if name not in seen:
                            seen.add(name)
                            all_names.append(name)
                    by_cat[label] = cat_list
                wb.close()
            except Exception:
                pass
        for name in ALL_METRICS:
            if name not in seen:
                seen.add(name)
                all_names.append(name)
        by_cat["Semua"] = all_names
        return by_cat, all_names

    def _update_paths(self):
        b = self.var_base_dir.get()
        self.var_companies_file.set(os.path.join(b, "companies.txt"))
        self.var_download_dir.set(os.path.join(b, "FinancialStatements"))
        self.var_xbrl_dir.set(os.path.join(b, "Folder_XBRL"))
        self.var_extract_dir.set(os.path.join(b, "ExtractedData_XBRL"))

    # ═════════════════════════════════════════════════════════════════
    #  BUILD UI
    # ═════════════════════════════════════════════════════════════════
    def _build_ui(self):
        # ── Sidebar ──
        self._sidebar = ctk.CTkFrame(self, fg_color=C_SIDEBAR, corner_radius=0, width=220)
        self._sidebar.pack(fill="y", side="left")
        self._sidebar.pack_propagate(False)

        # Garis aksen tipis di kanan sidebar
        ctk.CTkFrame(self._sidebar, fg_color=C_BORDER, width=1,
                     corner_radius=0).pack(side="right", fill="y")

        # Logo area
        logo = ctk.CTkFrame(self._sidebar, fg_color="transparent")
        logo.pack(fill="x", padx=20, pady=(28, 20))

        # Badge logo
        logo_badge = ctk.CTkFrame(logo, fg_color=C_ACCENT, corner_radius=10,
                                   width=40, height=40)
        logo_badge.pack(anchor="w")
        logo_badge.pack_propagate(False)
        ctk.CTkLabel(logo_badge, text="📊",
                     font=ctk.CTkFont("Segoe UI", 18)).place(relx=0.5, rely=0.5, anchor="center")

        ctk.CTkLabel(logo, text="IDX Superapp",
                     font=ctk.CTkFont("Segoe UI", 14, "bold"),
                     text_color=C_TEXT).pack(anchor="w", pady=(10, 0))
        ctk.CTkLabel(logo, text="Data Keuangan XBRL",
                     font=ctk.CTkFont("Segoe UI", 9),
                     text_color=C_MUTED).pack(anchor="w")

        ctk.CTkFrame(self._sidebar, fg_color=C_BORDER, height=1).pack(fill="x", padx=16, pady=(0, 12))

        # Nav label
        ctk.CTkLabel(self._sidebar, text="NAVIGASI",
                     font=ctk.CTkFont("Segoe UI", 9, "bold"),
                     text_color=C_MUTED).pack(anchor="w", padx=20, pady=(0, 6))

        # Nav items
        self._nav_pages = [
            "⚙️  Pengaturan",
            "📋  Pilih Akun",
            "🚀  Jalankan",
            "ℹ️  Tentang",
        ]
        self._nav_meta = {
            "⚙️  Pengaturan": ("⚙️", "Pengaturan"),
            "📋  Pilih Akun": ("📋", "Pilih Akun"),
            "🚀  Jalankan":   ("🚀", "Jalankan"),
            "ℹ️  Tentang":    ("ℹ️", "Tentang"),
        }
        self._nav_btns = {}
        for page in self._nav_pages:
            icon, label = self._nav_meta[page]
            btn = ctk.CTkButton(
                self._sidebar, text=f"  {icon}   {label}", anchor="w",
                font=ctk.CTkFont("Segoe UI", 12),
                fg_color="transparent",
                hover_color=C_CARD,
                text_color=C_MUTED2,
                height=42, corner_radius=8,
                command=lambda p=page: self._show_page(p)
            )
            btn.pack(fill="x", padx=10, pady=2)
            self._nav_btns[page] = btn

        # Sidebar bawah
        ctk.CTkFrame(self._sidebar, fg_color=C_BORDER, height=1).pack(
            fill="x", padx=16, side="bottom", pady=(0, 12))
        ctk.CTkLabel(self._sidebar, text="v1.0  ·  BEI / IDX",
                     font=ctk.CTkFont("Segoe UI", 8),
                     text_color=C_MUTED).pack(side="bottom", pady=(0, 4))

        # ── Content area ──
        self._content = ctk.CTkFrame(self, fg_color=C_BG, corner_radius=0)
        self._content.pack(fill="both", expand=True, side="left")

        self._pages = {}
        self._build_settings()
        self._build_account_selector()
        self._build_run()
        self._build_about()

        self._show_page(self._nav_pages[0])

    def _show_page(self, page: str):
        for frame in self._pages.values():
            frame.pack_forget()
        if page in self._pages:
            self._pages[page].pack(fill="both", expand=True)
        for p, btn in self._nav_btns.items():
            if p == page:
                btn.configure(fg_color=C_CARD, text_color=C_ACCENT_LT)
            else:
                btn.configure(fg_color="transparent", text_color=C_MUTED2)

    # ═════════════════════════════════════════════════════════════════
    #  PAGE 1 — PENGATURAN
    # ═════════════════════════════════════════════════════════════════
    def _build_settings(self):
        page_key = "⚙️  Pengaturan"
        tab = ctk.CTkFrame(self._content, fg_color=C_BG, corner_radius=0)
        self._pages[page_key] = tab

        # Page header
        self._page_header(tab, "⚙️  Pengaturan",
                          "Atur tahun laporan, periode, folder kerja, dan kurs sebelum menjalankan pipeline.")

        scroll = ctk.CTkScrollableFrame(tab, fg_color=C_BG, corner_radius=0,
                                        scrollbar_button_color=C_BORDER)
        scroll.pack(fill="both", expand=True, padx=24, pady=(0, 16))
        scroll.columnconfigure(0, weight=1)

        # ── Langkah 1 ──────────────────────────────────────────────
        self._section_header(scroll, "1", "Pilih Tahun & Periode Laporan",
                              "Pilih tahun fiskal dan jenis laporan yang ingin diunduh dari IDX.")

        card1 = self._card(scroll)
        # Tahun
        yr_row = ctk.CTkFrame(card1, fg_color="transparent")
        yr_row.pack(fill="x", pady=(0, 14))

        ctk.CTkLabel(yr_row, text="Tahun Laporan",
                     font=ctk.CTkFont("Segoe UI", 12, "bold"),
                     text_color=C_TEXT, anchor="w").pack(anchor="w", pady=(0, 6))

        yr_ctrl = ctk.CTkFrame(yr_row, fg_color="transparent")
        yr_ctrl.pack(anchor="w")
        ctk.CTkButton(yr_ctrl, text="◀", width=38, height=38,
                      fg_color=C_CARD2, hover_color=C_BORDER,
                      font=ctk.CTkFont("Segoe UI", 12, "bold"),
                      corner_radius=8,
                      command=lambda: self.var_year.set(self.var_year.get() - 1)).pack(side="left")
        ctk.CTkLabel(yr_ctrl, textvariable=self.var_year,
                     font=ctk.CTkFont("Segoe UI", 20, "bold"),
                     text_color=C_ACCENT, width=80).pack(side="left")
        ctk.CTkButton(yr_ctrl, text="▶", width=38, height=38,
                      fg_color=C_CARD2, hover_color=C_BORDER,
                      font=ctk.CTkFont("Segoe UI", 12, "bold"),
                      corner_radius=8,
                      command=lambda: self.var_year.set(self.var_year.get() + 1)).pack(side="left")

        # Divider
        ctk.CTkFrame(card1, fg_color=C_BORDER, height=1).pack(fill="x", pady=10)

        # Periode
        pd_row = ctk.CTkFrame(card1, fg_color="transparent")
        pd_row.pack(fill="x")
        ctk.CTkLabel(pd_row, text="Periode Laporan",
                     font=ctk.CTkFont("Segoe UI", 12, "bold"),
                     text_color=C_TEXT, anchor="w").pack(anchor="w", pady=(0, 6))

        pd_ctrl = ctk.CTkFrame(pd_row, fg_color="transparent")
        pd_ctrl.pack(fill="x")
        ctk.CTkComboBox(pd_ctrl, variable=self.var_period,
                        values=list(PERIOD_MAP.keys()),
                        state="readonly", width=280, height=38,
                        font=ctk.CTkFont("Segoe UI", 12),
                        dropdown_font=ctk.CTkFont("Segoe UI", 11),
                        fg_color=C_CARD2, border_color=C_BORDER,
                        button_color=C_ACCENT, button_hover_color=C_ACCENT2).pack(side="left")

        self._hint(pd_row, "Q1 = Maret  ·  Q2 = Juni  ·  Q3 = September  ·  Tahunan = Desember (Audit)")

        # ── Langkah 2 ──────────────────────────────────────────────
        self._section_header(scroll, "2", "Tentukan Folder Kerja",
                              "Semua file unduhan dan hasil ekstraksi akan disimpan di folder ini.")

        card2 = self._card(scroll)
        dir_row = ctk.CTkFrame(card2, fg_color="transparent")
        dir_row.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(dir_row, text="Folder Kerja",
                     font=ctk.CTkFont("Segoe UI", 12, "bold"),
                     text_color=C_TEXT).pack(anchor="w", pady=(0, 6))
        dir_ctrl = ctk.CTkFrame(dir_row, fg_color="transparent")
        dir_ctrl.pack(fill="x")
        ctk.CTkEntry(dir_ctrl, textvariable=self.var_base_dir,
                     font=ctk.CTkFont("Consolas", 10),
                     fg_color=C_CARD2, border_color=C_BORDER,
                     height=38, corner_radius=8).pack(side="left", fill="x", expand=True, padx=(0, 8))
        ctk.CTkButton(dir_ctrl, text="📁  Browse", width=110, height=38,
                      fg_color=C_ACCENT, hover_color=C_ACCENT2,
                      font=ctk.CTkFont("Segoe UI", 11, "bold"),
                      corner_radius=8, command=self._browse).pack(side="left")

        # Subfolder preview
        sub_frame = ctk.CTkFrame(card2, fg_color=C_CARD2, corner_radius=8)
        sub_frame.pack(fill="x", pady=(4, 0))
        sub_inner = ctk.CTkFrame(sub_frame, fg_color="transparent")
        sub_inner.pack(fill="x", padx=14, pady=10)
        ctk.CTkLabel(sub_inner, text="Struktur folder yang akan dibuat otomatis:",
                     font=ctk.CTkFont("Segoe UI", 9, "bold"),
                     text_color=C_MUTED2).pack(anchor="w", pady=(0, 6))
        for icon, lbl, var, desc in [
            ("📄", "companies.txt",       self.var_companies_file, "Daftar link download emiten"),
            ("📦", "FinancialStatements/", self.var_download_dir,  "File ZIP hasil download"),
            ("📂", "Folder_XBRL/",        self.var_xbrl_dir,       "File XBRL yang sudah diekstrak"),
            ("📊", "ExtractedData_XBRL/", self.var_extract_dir,    "Output akhir CSV + Excel"),
        ]:
            r = ctk.CTkFrame(sub_inner, fg_color="transparent")
            r.pack(fill="x", pady=2)
            ctk.CTkLabel(r, text=f"{icon}  {lbl}",
                         font=ctk.CTkFont("Consolas", 9),
                         text_color=C_ACCENT, width=200, anchor="w").pack(side="left")
            ctk.CTkLabel(r, text=f"— {desc}",
                         font=ctk.CTkFont("Segoe UI", 9),
                         text_color=C_MUTED).pack(side="left")

        # ── Langkah 3 ──────────────────────────────────────────────
        self._section_header(scroll, "3", "Kurs Mata Uang (USD → IDR)",
                              "Digunakan untuk mengonversi nilai keuangan yang dilaporkan dalam USD ke Rupiah.")

        card3 = self._card(scroll)
        ctk.CTkLabel(card3, text="Kurs USD ke IDR",
                     font=ctk.CTkFont("Segoe UI", 12, "bold"),
                     text_color=C_TEXT).pack(anchor="w", pady=(0, 6))
        usd_ctrl = ctk.CTkFrame(card3, fg_color="transparent")
        usd_ctrl.pack(anchor="w")
        ctk.CTkEntry(usd_ctrl, textvariable=self.var_usd_rate,
                     width=200, height=38,
                     font=ctk.CTkFont("Consolas", 13),
                     fg_color=C_CARD2, border_color=C_BORDER,
                     corner_radius=8).pack(side="left")
        ctk.CTkLabel(usd_ctrl, text="IDR per 1 USD",
                     font=ctk.CTkFont("Segoe UI", 10),
                     text_color=C_MUTED2).pack(side="left", padx=12)
        self._hint(card3, "Contoh: 16000 berarti 1 USD = Rp 16.000. Hanya mempengaruhi emiten yang melaporkan dalam USD.")

        # Tombol simpan (visual saja, pengaturan otomatis tersimpan di var)
        ctk.CTkButton(scroll, text="✓  Lanjut ke Pilih Akun →",
                      fg_color=C_ACCENT, hover_color=C_ACCENT2,
                      font=ctk.CTkFont("Segoe UI", 12, "bold"),
                      height=44, corner_radius=10,
                      command=lambda: self._show_page("📋  Pilih Akun")).pack(fill="x", pady=(8, 4))

    def _browse(self):
        d = filedialog.askdirectory(title="Pilih Folder Kerja")
        if d:
            self.var_base_dir.set(d.replace("/", "\\"))

    # ═════════════════════════════════════════════════════════════════
    #  PAGE 2 — PILIH AKUN
    # ═════════════════════════════════════════════════════════════════
    def _build_account_selector(self):
        page_key = "📋  Pilih Akun"
        tab = ctk.CTkFrame(self._content, fg_color=C_BG, corner_radius=0)
        self._pages[page_key] = tab

        self._page_header(tab, "📋  Pilih Akun",
                          "Pilih akun keuangan yang ingin diekstrak. Gunakan preset untuk analisis Altman Z-Score atau IBD.")

        # ── 2-column body ──
        body = ctk.CTkFrame(tab, fg_color=C_BG, corner_radius=0)
        body.pack(fill="both", expand=True, padx=16, pady=(0, 16))
        body.columnconfigure(0, weight=5)
        body.columnconfigure(1, weight=4)
        body.rowconfigure(0, weight=1)

        # ═══ KOLOM KIRI — Cari & Tambah ═══
        left_col = ctk.CTkFrame(body, fg_color=C_SURFACE, corner_radius=12,
                                border_color=C_BORDER, border_width=1)
        left_col.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        left_col.rowconfigure(3, weight=1)
        left_col.columnconfigure(0, weight=1)

        # Header kiri
        hdr_l = ctk.CTkFrame(left_col, fg_color="transparent")
        hdr_l.grid(row=0, column=0, sticky="ew", padx=16, pady=(14, 10))
        ctk.CTkLabel(hdr_l, text="🔍  Cari Akun",
                     font=ctk.CTkFont("Segoe UI", 13, "bold"),
                     text_color=C_WHITE).pack(side="left")
        total_count = len(self._all_account_options)
        ctk.CTkLabel(hdr_l, text=f"{total_count:,} akun tersedia",
                     font=ctk.CTkFont("Segoe UI", 9),
                     text_color=C_MUTED,
                     fg_color=C_CARD2, corner_radius=6,
                     padx=8, pady=3).pack(side="right")

        # Filter kategori
        cat_wrap = ctk.CTkFrame(left_col, fg_color="transparent")
        cat_wrap.grid(row=1, column=0, sticky="ew", padx=16, pady=(0, 8))
        cats = ["Semua", "Informasi Umum", "Neraca", "Laba Rugi", "Arus Kas"]
        cat_short = {
            "Semua": "Semua",
            "Informasi Umum": "Info Umum",
            "Neraca": "Neraca",
            "Laba Rugi": "Laba Rugi",
            "Arus Kas": "Arus Kas",
        }
        cat_counts = {c: len(self._accounts_by_category.get(c, [])) for c in cats}
        for cat in cats:
            is_active = (cat == self._active_category)
            n = cat_counts.get(cat, 0)
            label_text = f"{cat_short[cat]}  {n}" if cat != "Semua" else "Semua"
            btn = ctk.CTkButton(
                cat_wrap, text=label_text,
                height=30, font=ctk.CTkFont("Segoe UI", 9, "bold"),
                corner_radius=8,
                fg_color=C_ACCENT if is_active else C_CARD2,
                hover_color=C_ACCENT2,
                text_color=C_WHITE if is_active else C_MUTED2,
                command=lambda c=cat: self._set_category(c),
            )
            btn.pack(side="left", padx=(0, 5))
            self._cat_buttons[cat] = btn

        # Search bar
        search_row = ctk.CTkFrame(left_col, fg_color="transparent")
        search_row.grid(row=2, column=0, sticky="ew", padx=16, pady=(0, 8))
        ctk.CTkEntry(search_row, textvariable=self._search_var,
                     placeholder_text="🔍  Ketik nama akun... (contoh: aset, laba, kas)",
                     font=ctk.CTkFont("Segoe UI", 11),
                     fg_color=C_CARD2, border_color=C_BORDER,
                     height=40, corner_radius=8).pack(side="left", fill="x", expand=True)
        ctk.CTkButton(search_row, text="✕", width=40, height=40,
                      fg_color=C_CARD2, hover_color=C_BORDER,
                      font=ctk.CTkFont("Segoe UI", 12, "bold"),
                      corner_radius=8,
                      command=lambda: self._search_var.set("")).pack(side="left", padx=(6, 0))

        # Listbox
        lb_outer = ctk.CTkFrame(left_col, fg_color=C_CARD2, corner_radius=8,
                                border_color=C_BORDER, border_width=1)
        lb_outer.grid(row=3, column=0, sticky="nsew", padx=16, pady=(0, 4))
        self.account_listbox = tk.Listbox(
            lb_outer,
            selectmode=tk.SINGLE,
            bg=C_CARD2, fg=C_TEXT,
            selectbackground=C_ACCENT, selectforeground=C_WHITE,
            highlightthickness=0, relief="flat", activestyle="none",
            font=("Segoe UI", 10), cursor="hand2",
        )
        self.account_listbox.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=8)
        lb_sc = tk.Scrollbar(lb_outer, orient="vertical", command=self.account_listbox.yview)
        lb_sc.pack(side="right", fill="y", padx=(4, 6), pady=8)
        self.account_listbox.configure(yscrollcommand=lb_sc.set)
        self.account_listbox.bind("<Double-Button-1>", lambda _e: self._add_selected_metric())
        self.account_listbox.bind("<MouseWheel>", self._on_account_list_scroll)
        self.account_listbox.bind("<Button-4>", self._on_account_list_scroll)
        self.account_listbox.bind("<Button-5>", self._on_account_list_scroll)

        # Hint
        hint_lb = ctk.CTkFrame(left_col, fg_color="transparent")
        hint_lb.grid(row=4, column=0, sticky="ew", padx=16, pady=(2, 4))
        ctk.CTkLabel(hint_lb, text="💡  Klik dua kali atau tekan tombol di bawah untuk menambahkan akun",
                     font=ctk.CTkFont("Segoe UI", 9), text_color=C_MUTED).pack(anchor="w")

        # Tombol tambah
        ctk.CTkButton(left_col, text="＋  Tambahkan Akun yang Dipilih",
                      fg_color=C_ACCENT, hover_color=C_ACCENT2,
                      font=ctk.CTkFont("Segoe UI", 11, "bold"),
                      height=42, corner_radius=8,
                      command=self._add_selected_metric).grid(
                          row=5, column=0, sticky="ew", padx=16, pady=(0, 14))

        # ═══ KOLOM KANAN — Preset + Akun Terpilih ═══
        right_col = ctk.CTkFrame(body, fg_color=C_BG, corner_radius=0)
        right_col.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
        right_col.rowconfigure(1, weight=1)
        right_col.columnconfigure(0, weight=1)

        # ── Preset cards ──
        preset_outer = ctk.CTkFrame(right_col, fg_color=C_SURFACE, corner_radius=12,
                                    border_color=C_BORDER, border_width=1)
        preset_outer.grid(row=0, column=0, sticky="ew", pady=(0, 8))

        ph = ctk.CTkFrame(preset_outer, fg_color="transparent")
        ph.pack(fill="x", padx=16, pady=(14, 10))
        ctk.CTkLabel(ph, text="⚗️  Preset Analisis Keuangan",
                     font=ctk.CTkFont("Segoe UI", 13, "bold"),
                     text_color=C_WHITE).pack(side="left")
        ctk.CTkLabel(ph, text="Sekali klik — otomatis pilih semua akun",
                     font=ctk.CTkFont("Segoe UI", 9), text_color=C_MUTED).pack(side="right")

        pr = ctk.CTkFrame(preset_outer, fg_color="transparent")
        pr.pack(fill="x", padx=14, pady=(0, 14))
        pr.columnconfigure(0, weight=1)
        pr.columnconfigure(1, weight=1)

        # Kartu Altman
        af = ctk.CTkFrame(pr, fg_color="#13111c", corner_radius=10,
                          border_color="#4c1d95", border_width=1)
        af.grid(row=0, column=0, sticky="ew", padx=(0, 6))
        ai = ctk.CTkFrame(af, fg_color="transparent")
        ai.pack(fill="x", padx=14, pady=14)
        ctk.CTkLabel(ai, text="Altman Z-Score",
                     font=ctk.CTkFont("Segoe UI", 12, "bold"),
                     text_color=C_PURPLE).pack(anchor="w")
        ctk.CTkLabel(ai, text=f"{len(self.ALTMAN_ACCOUNTS)} akun",
                     font=ctk.CTkFont("Segoe UI", 9),
                     text_color=C_PURPLE,
                     fg_color="#2e1065", corner_radius=4,
                     padx=7, pady=2).pack(anchor="w", pady=(4, 0))
        ctk.CTkLabel(ai, text="Prediksi risiko kebangkrutan\nberdasarkan rasio keuangan.",
                     font=ctk.CTkFont("Segoe UI", 9), text_color=C_MUTED,
                     justify="left").pack(anchor="w", pady=(6, 10))
        ctk.CTkButton(ai, text="Pilih Akun Altman",
                      fg_color="#4c1d95", hover_color="#5b21b6",
                      text_color=C_PURPLE,
                      font=ctk.CTkFont("Segoe UI", 10, "bold"),
                      height=34, corner_radius=8,
                      command=self._select_altman_accounts).pack(fill="x")

        # Kartu IBD
        ibdf = ctk.CTkFrame(pr, fg_color="#0a1a12", corner_radius=10,
                             border_color="#065f46", border_width=1)
        ibdf.grid(row=0, column=1, sticky="ew", padx=(6, 0))
        ii = ctk.CTkFrame(ibdf, fg_color="transparent")
        ii.pack(fill="x", padx=14, pady=14)
        ctk.CTkLabel(ii, text="IBD",
                     font=ctk.CTkFont("Segoe UI", 12, "bold"),
                     text_color=C_GREEN).pack(anchor="w")
        ctk.CTkLabel(ii, text=f"{len(self.IBD_ACCOUNTS)} akun",
                     font=ctk.CTkFont("Segoe UI", 9),
                     text_color=C_GREEN,
                     fg_color="#064e3b", corner_radius=4,
                     padx=7, pady=2).pack(anchor="w", pady=(4, 0))
        ctk.CTkLabel(ii, text="Interest Bearing Debt —\nutang berbunga dari neraca.",
                     font=ctk.CTkFont("Segoe UI", 9), text_color=C_MUTED,
                     justify="left").pack(anchor="w", pady=(6, 10))
        ctk.CTkButton(ii, text="Pilih Akun IBD",
                      fg_color="#065f46", hover_color="#047857",
                      text_color=C_GREEN,
                      font=ctk.CTkFont("Segoe UI", 10, "bold"),
                      height=34, corner_radius=8,
                      command=self._select_ibd_accounts).pack(fill="x")

        # ── Akun Terpilih ──
        sel_outer = ctk.CTkFrame(right_col, fg_color=C_SURFACE, corner_radius=12,
                                 border_color=C_BORDER, border_width=1)
        sel_outer.grid(row=1, column=0, sticky="nsew")
        sel_outer.rowconfigure(1, weight=1)
        sel_outer.columnconfigure(0, weight=1)

        sel_hdr = ctk.CTkFrame(sel_outer, fg_color="transparent")
        sel_hdr.grid(row=0, column=0, sticky="ew", padx=16, pady=(14, 8))

        self.lbl_selected_count = ctk.CTkLabel(
            sel_hdr, text=self._count_text(),
            font=ctk.CTkFont("Segoe UI", 12, "bold"), text_color=C_WHITE)
        self.lbl_selected_count.pack(side="left")

        btn_kw = dict(height=28, font=ctk.CTkFont("Segoe UI", 9, "bold"),
                      corner_radius=7, fg_color=C_CARD2, hover_color=C_BORDER)
        ctk.CTkButton(sel_hdr, text="↺ Default", width=80, **btn_kw,
                      command=self._reset_to_default).pack(side="right", padx=(4, 0))
        ctk.CTkButton(sel_hdr, text="☐ Hapus Semua", width=110, **btn_kw,
                      text_color=C_WARNING,
                      command=self._deselect_all).pack(side="right", padx=(4, 0))

        self.selected_list = ctk.CTkScrollableFrame(
            sel_outer, fg_color=C_CARD2, corner_radius=8,
            scrollbar_button_color=C_BORDER,
        )
        self.selected_list.grid(row=1, column=0, sticky="nsew", padx=12, pady=(0, 12))

        # Lanjut button
        ctk.CTkButton(right_col, text="✓  Lanjut ke Jalankan →",
                      fg_color=C_SUCCESS, hover_color=C_SUCCESS2,
                      font=ctk.CTkFont("Segoe UI", 11, "bold"),
                      height=40, corner_radius=10,
                      command=lambda: self._show_page("🚀  Jalankan")).grid(
                          row=2, column=0, sticky="ew", pady=(4, 0))

        self._refresh_dropdown_options()
        self._refresh_selected_metrics_view()

    # ─────────────────────────────────────────────────────────────────
    def _count_text(self) -> str:
        n = len(self._selected_metrics)
        return f"✅  {n} Akun Terpilih"

    def _refresh_dropdown_options(self):
        if not self.account_listbox:
            return
        q = self._search_var.get().strip().lower()
        selected = set(self._selected_metrics)
        cat = self._active_category
        source = (self._accounts_by_category.get(cat, self._all_account_options)
                  if cat and cat != "Semua"
                  else self._all_account_options)
        self._filtered_metrics = [
            m for m in source
            if m not in selected and (not q or q in m.lower())
        ]
        self.account_listbox.delete(0, tk.END)
        if self._filtered_metrics:
            for m in self._filtered_metrics:
                self.account_listbox.insert(tk.END, m)
            self.account_listbox.selection_set(0)
        else:
            self.account_listbox.insert(tk.END, "(Tidak ada hasil — coba kata kunci lain)")

    def _refresh_selected_metrics_view(self):
        if not self.selected_list:
            return
        for child in self.selected_list.winfo_children():
            child.destroy()

        if not self._selected_metrics:
            empty = ctk.CTkFrame(self.selected_list, fg_color="transparent")
            empty.pack(fill="x", pady=20)
            ctk.CTkLabel(empty, text="Belum ada akun dipilih",
                         font=ctk.CTkFont("Segoe UI", 11, "bold"),
                         text_color=C_MUTED).pack()
            ctk.CTkLabel(empty, text="Cari dan tambahkan akun dari kolom kiri,\natau gunakan tombol preset di atas.",
                         font=ctk.CTkFont("Segoe UI", 9), text_color=C_MUTED,
                         justify="center").pack(pady=(4, 0))
        else:
            for i, metric in enumerate(self._selected_metrics):
                row = ctk.CTkFrame(self.selected_list,
                                   fg_color=C_SURFACE if i % 2 == 0 else C_CARD2,
                                   corner_radius=6)
                row.pack(fill="x", padx=4, pady=2)
                ctk.CTkLabel(row, text=f"  {i+1}.",
                             font=ctk.CTkFont("Consolas", 9),
                             text_color=C_MUTED, width=30, anchor="e").pack(side="left")
                ctk.CTkLabel(row, text=metric,
                             font=ctk.CTkFont("Segoe UI", 10),
                             text_color=C_TEXT, anchor="w").pack(
                                 side="left", fill="x", expand=True, padx=(6, 0), pady=6)
                ctk.CTkButton(row, text="✕", width=28, height=24,
                              fg_color="transparent", hover_color=C_DANGER,
                              text_color=C_MUTED, font=ctk.CTkFont("Segoe UI", 10, "bold"),
                              corner_radius=6,
                              command=lambda m=metric: self._remove_selected_metric(m)).pack(
                                  side="right", padx=6)

        if self.lbl_selected_count:
            self.lbl_selected_count.configure(text=self._count_text())

    def _add_selected_metric(self):
        if not self.account_listbox: return
        cur = self.account_listbox.curselection()
        if not cur: return
        metric = self.account_listbox.get(cur[0]).strip()
        if not metric or metric.startswith("("): return
        if metric not in self._selected_metrics:
            self._selected_metrics.append(metric)
        self._refresh_dropdown_options()
        self._refresh_selected_metrics_view()

    def _on_account_list_scroll(self, event):
        if not self.account_listbox: return "break"
        delta = 0
        if hasattr(event, "delta") and event.delta:
            delta = -1 if event.delta > 0 else 1
        elif getattr(event, "num", None) == 4:
            delta = -1
        elif getattr(event, "num", None) == 5:
            delta = 1
        if delta: self.account_listbox.yview_scroll(delta, "units")
        return "break"

    def _remove_selected_metric(self, metric: str):
        self._selected_metrics = [m for m in self._selected_metrics if m != metric]
        self._refresh_dropdown_options()
        self._refresh_selected_metrics_view()

    def _select_all(self):
        self._selected_metrics = list(self._all_account_options)
        self._refresh_dropdown_options()
        self._refresh_selected_metrics_view()

    def _deselect_all(self):
        self._selected_metrics = []
        self._refresh_dropdown_options()
        self._refresh_selected_metrics_view()

    def _reset_to_default(self):
        self._selected_metrics = [m for m in DEFAULT_METRICS if m in self._all_account_options]
        self._refresh_dropdown_options()
        self._refresh_selected_metrics_view()

    def _filter_metrics(self):
        self._refresh_dropdown_options()

    def _set_category(self, cat: str):
        self._active_category = cat
        for c, btn in self._cat_buttons.items():
            if c == cat:
                btn.configure(fg_color=C_ACCENT, text_color=C_WHITE)
            else:
                btn.configure(fg_color=C_CARD2, text_color=C_MUTED2)
        self._refresh_dropdown_options()

    # ─── Preset lists ───────────────────────────────────────────────
    _IDENTITY_ACCOUNTS = [
        "Nama entitas", "Kode entitas", "Sektor", "Subsektor", "Industri", "Subindustri",
        "Periode penyampaian laporan keuangan",
        "Mata uang pelaporan",
        "Pembulatan yang digunakan dalam penyajian jumlah dalam laporan keuangan",
    ]

    ALTMAN_ACCOUNTS = _IDENTITY_ACCOUNTS + [
        "Jumlah aset", "Jumlah aset lancar", "Jumlah liabilitas jangka pendek",
        "Jumlah ekuitas",
        "Saldo laba yang belum ditentukan penggunaannya",
        "Saldo laba yang telah ditentukan penggunaannya",
        "Penjualan dan pendapatan usaha",
        "Jumlah laba (rugi) sebelum pajak penghasilan",
        "Beban bunga dan keuangan", "Jumlah liabilitas",
        "Depresiasi", "Amortisasi",
        "Beban bunga", "Pendapatan bunga", "Pendapatan operasional lainnya",
    ]

    IBD_ACCOUNTS = _IDENTITY_ACCOUNTS + [
        "Utang bank jangka pendek", "Utang trust receipts",
        "Pinjaman jangka pendek non-bank", "Utang lembaga keuangan non-bank",
        "Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas utang bank",
        "Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas utang keuangan keuangan non bank",
        "Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas utang obligasi",
        "Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas surat utang jangka menengah",
        "Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas utang pembiayaan konsumen",
        "Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas liabilitas sewa pembiayaan",
        "Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas sukuk",
        "Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas obligasi subordinasi",
        "Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas pinjaman beragunan",
        "Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas pinjaman tanpa agunan",
        "Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas penerusan pinjaman",
        "Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas pinjaman dari pemerintah republik Indonesia",
        "Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas pinjaman subordinasi",
        "Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas liabilitas kerja sama operasi",
        "Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas liabilitas pembebasan tanah",
        "Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas utang listrik swasta",
        "Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas utang retensi",
        "Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas wesel bayar",
        "Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas pinjaman lainnya",
        "Liabilitas jangka panjang atas utang bank",
        "Liabilitas jangka panjang atas utang obligasi",
        "Liabilitas jangka panjang atas pinjaman beragunan",
        "Liabilitas jangka panjang atas pinjaman tanpa agunan",
        "Liabilitas jangka panjang atas pinjaman dari pemerintah republik Indonesia",
        "Liabilitas jangka panjang atas pinjaman subordinasi",
        "Liabilitas jangka panjang atas utang pembiayaan konsumen",
        "Liabilitas jangka panjang atas surat utang jangka menengah",
        "Liabilitas jangka panjang atas sukuk",
        "Liabilitas jangka panjang atas obligasi subordinasi",
        "Liabilitas jangka panjang atas liabilitas sewa pembiayaan",
        "Liabilitas jangka panjang atas penerusan pinjaman",
        "Liabilitas jangka panjang atas liabilitas kerja sama operasi",
        "Liabilitas jangka panjang atas liabilitas pembebasan tanah",
        "Liabilitas jangka panjang atas utang listrik swasta",
        "Liabilitas jangka panjang atas utang retensi",
        "Liabilitas jangka panjang atas wesel bayar",
        "Liabilitas jangka panjang atas pinjaman lainnya",
        "Utang obligasi", "Sukuk", "Obligasi subordinasi",
        "Liabilitas sewa pembiayaan", "Obligasi konversi",
        "Pinjaman subordinasi pihak ketiga", "Pinjaman subordinasi pihak berelasi",
        "Sukuk mudharabah", "Sukuk mudharabah subordinasi",
    ]

    def _select_altman_accounts(self):
        for acc in self.ALTMAN_ACCOUNTS:
            if acc in self._all_account_options and acc not in self._selected_metrics:
                self._selected_metrics.append(acc)
        self._refresh_dropdown_options()
        self._refresh_selected_metrics_view()

    def _select_ibd_accounts(self):
        for acc in self.IBD_ACCOUNTS:
            if acc in self._all_account_options and acc not in self._selected_metrics:
                self._selected_metrics.append(acc)
        self._refresh_dropdown_options()
        self._refresh_selected_metrics_view()

    def get_selected_metrics(self) -> list:
        return list(self._selected_metrics)

    # ═════════════════════════════════════════════════════════════════
    #  PAGE 3 — JALANKAN
    # ═════════════════════════════════════════════════════════════════
    def _build_run(self):
        page_key = "🚀  Jalankan"
        tab = ctk.CTkFrame(self._content, fg_color=C_BG, corner_radius=0)
        self._pages[page_key] = tab

        self._page_header(tab, "🚀  Jalankan Pipeline",
                          "Klik 'Jalankan Semua' untuk memulai proses ekstraksi secara otomatis dari awal hingga akhir.")

        # ── Root: PanedWindow horizontal — fixed split, TIDAK re-layout saat update ──
        paned = tk.PanedWindow(tab, orient=tk.HORIZONTAL,
                               bg=C_BG, sashwidth=6, sashrelief="flat",
                               bd=0, relief="flat")
        paned.pack(fill="both", expand=True, padx=16, pady=(0, 16))

        # ═══ PANEL KIRI — Pipeline (tk native, zero CTk re-layout) ═══
        left_root = tk.Frame(paned, bg=C_BG)
        paned.add(left_root, minsize=380, stretch="always")

        # Container pipeline (border simulasi dengan frame)
        pipe_border = tk.Frame(left_root, bg=C_BORDER)
        pipe_border.pack(fill="both", expand=True, pady=(0, 8))
        pipe_bg = tk.Frame(pipe_border, bg=C_SURFACE)
        pipe_bg.pack(fill="both", expand=True, padx=1, pady=1)

        # Header pipeline (tk native)
        pipe_hdr = tk.Frame(pipe_bg, bg=C_SURFACE)
        pipe_hdr.pack(fill="x", padx=16, pady=(14, 10))
        tk.Label(pipe_hdr, text="🔄  Pipeline Ekstraksi Data",
                 font=("Segoe UI", 12, "bold"),
                 fg=C_WHITE, bg=C_SURFACE, anchor="w").pack(side="left")
        tk.Label(pipe_hdr, text=" 7 langkah otomatis ",
                 font=("Segoe UI", 9),
                 fg=C_MUTED, bg=C_CARD2).pack(side="right")

        # Scroll area pipeline — tk.Canvas + tk.Scrollbar (100% native, nol re-layout)
        pipe_canvas = tk.Canvas(pipe_bg, bg=C_SURFACE, highlightthickness=0, bd=0)
        pipe_scrollbar = tk.Scrollbar(pipe_bg, orient="vertical",
                                      command=pipe_canvas.yview)
        pipe_canvas.configure(yscrollcommand=pipe_scrollbar.set)
        pipe_scrollbar.pack(side="right", fill="y", pady=(0, 8))
        pipe_canvas.pack(side="left", fill="both", expand=True, padx=(0, 0))

        # Frame di dalam canvas
        pipe_inner = tk.Frame(pipe_canvas, bg=C_SURFACE)
        pipe_canvas_window = pipe_canvas.create_window((0, 0), window=pipe_inner,
                                                        anchor="nw", tags="inner")

        def _on_pipe_configure(e):
            pipe_canvas.configure(scrollregion=pipe_canvas.bbox("all"))
        def _on_canvas_resize(e):
            pipe_canvas.itemconfig(pipe_canvas_window, width=e.width)
        pipe_inner.bind("<Configure>", _on_pipe_configure)
        pipe_canvas.bind("<Configure>", _on_canvas_resize)
        pipe_canvas.bind("<MouseWheel>",
                         lambda e: pipe_canvas.yview_scroll(-1 if e.delta > 0 else 1, "units"))

        self.steps = [
            ("generate", "Generate Links",     "🔗", "1", "Mengambil semua ticker dari IDX dan membuat daftar link download"),
            ("download",  "Download XBRL",     "⬇",  "2", "Mengunduh file ZIP laporan keuangan dari IDX.co.id"),
            ("unzip",     "Ekstrak ZIP",        "📦", "3", "Membuka file ZIP dan mengekstrak isinya ke folder XBRL"),
            ("extract",   "Ekstrak Data XBRL", "🔍", "4", "Membaca file XBRL dan mengekstrak akun yang dipilih ke CSV"),
            ("normalize", "Normalisasi IDR",   "💱", "5", "Mengonversi semua nilai ke Rupiah penuh berdasarkan kurs"),
            ("export",    "Export Excel",      "📊", "6", "Menyimpan hasil ke file Excel (sebelum & sesudah konversi)"),
            ("metrics",   "Hitung Metrik",     "⚗", "7", "Menghitung Altman Z-Score dan IBD dari data yang sudah diekspor"),
        ]
        self._step_w = {}
        for key, label, icon, num, desc in self.steps:
            self._make_step_row(pipe_inner, key, label, icon, num, desc)
        # Spacer bawah
        tk.Frame(pipe_inner, bg=C_SURFACE, height=8).pack(fill="x")

        # Tombol Run/Stop (tk native frame, CTkButton masih ok karena tidak di dalam canvas)
        btn_border = tk.Frame(left_root, bg=C_BORDER)
        btn_border.pack(fill="x")
        btn_bg = tk.Frame(btn_border, bg=C_SURFACE)
        btn_bg.pack(fill="x", padx=1, pady=1)
        btn_inner_f = tk.Frame(btn_bg, bg=C_SURFACE)
        btn_inner_f.pack(fill="x", padx=14, pady=12)

        self.btn_run_all = ctk.CTkButton(
            btn_inner_f, text="▶  Jalankan Semua (8 Langkah)",
            font=ctk.CTkFont("Segoe UI", 13, "bold"),
            fg_color=C_SUCCESS, hover_color=C_SUCCESS2,
            height=50, corner_radius=10,
            command=self._run_all)
        self.btn_run_all.pack(side="left", fill="x", expand=True, padx=(0, 8))

        self.btn_stop = ctk.CTkButton(
            btn_inner_f, text="⏹  Stop",
            font=ctk.CTkFont("Segoe UI", 12, "bold"),
            fg_color=C_CARD2, hover_color=C_DANGER,
            text_color=C_MUTED2,
            height=50, width=120, corner_radius=10,
            state="disabled", command=self._stop_pipeline)
        self.btn_stop.pack(side="left")

        # ═══ PANEL KANAN — Progress + Log ═══
        right_root = tk.Frame(paned, bg=C_BG)
        paned.add(right_root, minsize=320, stretch="always")

        # Progress card — tk native, overall_pb pakai tk.Canvas
        prog_border = tk.Frame(right_root, bg=C_BORDER)
        prog_border.pack(fill="x", pady=(0, 8))
        prog_bg = tk.Frame(prog_border, bg=C_SURFACE)
        prog_bg.pack(fill="x", padx=1, pady=1)
        prog_inner = tk.Frame(prog_bg, bg=C_SURFACE)
        prog_inner.pack(fill="x", padx=16, pady=14)

        prog_row = tk.Frame(prog_inner, bg=C_SURFACE)
        prog_row.pack(fill="x", pady=(0, 8))
        tk.Label(prog_row, text="📈  Progress Keseluruhan",
                 font=("Segoe UI", 12, "bold"),
                 fg=C_WHITE, bg=C_SURFACE).pack(side="left")
        self.lbl_overall_pct = tk.Label(prog_row, text="0%",
                                         font=("Segoe UI", 12, "bold"),
                                         fg=C_ACCENT, bg=C_SURFACE)
        self.lbl_overall_pct.pack(side="right")

        # Overall progress: tk.Canvas — tidak ada CTk overhead sama sekali
        self.overall_pb = tk.Canvas(prog_inner, height=8, bg=C_CARD2,
                                     highlightthickness=0, bd=0)
        self.overall_pb.pack(fill="x")
        self.overall_pb.create_rectangle(0, 0, 0, 8, fill=C_ACCENT,
                                          outline="", tags="fill")
        # Simpan lebar terakhir untuk kalkulasi fill
        self._overall_pb_w = 0
        self.overall_pb.bind("<Configure>",
                              lambda e: setattr(self, "_overall_pb_w", e.width))

        self.lbl_overall = tk.Label(prog_inner,
                                     text="Siap — tekan Jalankan Semua untuk memulai",
                                     font=("Segoe UI", 10),
                                     fg=C_MUTED2, bg=C_SURFACE,
                                     anchor="w", justify="left")
        self.lbl_overall.pack(fill="x", pady=(8, 0))

        # Log card
        log_border = tk.Frame(right_root, bg=C_BORDER)
        log_border.pack(fill="both", expand=True)
        log_bg = tk.Frame(log_border, bg=C_SURFACE)
        log_bg.pack(fill="both", expand=True, padx=1, pady=1)

        log_hdr = tk.Frame(log_bg, bg=C_SURFACE)
        log_hdr.pack(fill="x", padx=16, pady=(12, 6))
        tk.Label(log_hdr, text="📃  Log Aktivitas",
                 font=("Segoe UI", 12, "bold"),
                 fg=C_WHITE, bg=C_SURFACE).pack(side="left")
        ctk.CTkButton(log_hdr, text="🗑  Hapus", width=80, height=28,
                      fg_color=C_CARD2, hover_color=C_BORDER,
                      font=ctk.CTkFont("Segoe UI", 9),
                      corner_radius=7,
                      command=self._clear_log).pack(side="right")

        self.log_box = ctk.CTkTextbox(
            log_bg, fg_color=C_CARD2, text_color=C_MUTED2,
            font=ctk.CTkFont("Consolas", 10), wrap="word",
            corner_radius=8, border_color=C_BORDER, border_width=1,
            scrollbar_button_color=C_BORDER2, state="disabled"
        )
        self.log_box.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        tb = self.log_box._textbox
        tb.tag_configure("ok",   foreground=C_SUCCESS)
        tb.tag_configure("err",  foreground=C_DANGER)
        tb.tag_configure("warn", foreground=C_WARNING)
        tb.tag_configure("info", foreground=C_ACCENT_LT)
        tb.tag_configure("dim",  foreground=C_MUTED)

    def _make_step_row(self, parent, key, label, icon, num, desc):
        # Seluruhnya tk native — nol CTk widget di dalam pipeline scroll
        card = tk.Frame(parent, bg=C_CARD)
        card.pack(fill="x", padx=8, pady=3)

        inner = tk.Frame(card, bg=C_CARD)
        inner.pack(fill="x", padx=14, pady=11)

        # Badge nomor — indigo pill
        num_badge = tk.Label(inner, text=f" {num} ",
                              font=("Segoe UI", 9, "bold"),
                              fg=C_ACCENT_LT, bg="#1e1b4b",
                              relief="flat", bd=0)
        num_badge.pack(side="left", padx=(0, 12))

        # Info kiri
        info = tk.Frame(inner, bg=C_CARD)
        info.pack(side="left", fill="x", expand=True)
        tk.Label(info, text=f"{icon}  {label}",
                 font=("Segoe UI", 11, "bold"),
                 fg=C_TEXT, bg=C_CARD, anchor="w").pack(anchor="w")
        tk.Label(info, text=desc,
                 font=("Segoe UI", 8),
                 fg=C_MUTED, bg=C_CARD, anchor="w").pack(anchor="w")

        # Kontrol kanan — semua tk native
        ctrl = tk.Frame(inner, bg=C_CARD)
        ctrl.pack(side="right", padx=(10, 0))

        # Progress bar: tk.Canvas — track lebih ramping & elegan
        pb_canvas = tk.Canvas(ctrl, width=130, height=6,
                              bg=C_CARD, highlightthickness=0, bd=0)
        pb_canvas.pack(side="left", padx=(0, 10))
        pb_canvas.create_rectangle(0, 0, 130, 6, fill=C_BORDER2, outline="", tags="track")
        pb_canvas.create_rectangle(0, 0, 0,   6, fill=C_ACCENT,  outline="", tags="fill")

        # Label status: tk.Label
        lbl = tk.Label(ctrl, text="Menunggu",
                       font=("Segoe UI", 9),
                       fg=C_MUTED, bg=C_CARD,
                       width=9, anchor="w", bd=0)
        lbl.pack(side="left", padx=(0, 8))

        # Tombol run individual: tk.Button
        btn = tk.Button(ctrl, text="▶",
                        font=("Segoe UI", 9, "bold"),
                        fg=C_ACCENT_LT, bg=C_CARD2,
                        activebackground=C_ACCENT, activeforeground=C_WHITE,
                        relief="flat", bd=0, cursor="hand2", width=3,
                        command=lambda k=key: self._run_step(k))
        btn.pack(side="left")

        self._step_w[key] = {"pb": pb_canvas, "lbl": lbl, "btn": btn, "num": num_badge}

    # ═════════════════════════════════════════════════════════════════
    #  PAGE 4 — TENTANG
    # ═════════════════════════════════════════════════════════════════
    def _build_about(self):
        page_key = "ℹ️  Tentang"
        tab = ctk.CTkFrame(self._content, fg_color=C_BG, corner_radius=0)
        self._pages[page_key] = tab

        self._page_header(tab, "ℹ️  Tentang IDX Superapp",
                          "Alat otomatis untuk mengunduh dan mengekstrak data keuangan dari Bursa Efek Indonesia.")

        scroll = ctk.CTkScrollableFrame(tab, fg_color=C_BG, corner_radius=0,
                                        scrollbar_button_color=C_BORDER)
        scroll.pack(fill="both", expand=True, padx=24, pady=(0, 16))

        # Panduan singkat
        self._section_header(scroll, "📖", "Cara Menggunakan",
                              "Ikuti 3 langkah sederhana ini untuk mendapatkan data keuangan IDX.")

        guide_card = self._card(scroll)
        steps_guide = [
            ("1️⃣", "Atur Pengaturan",
             "Buka halaman Pengaturan → tentukan tahun, periode, dan folder kerja."),
            ("2️⃣", "Pilih Akun",
             "Buka halaman Pilih Akun → gunakan preset Altman / IBD, atau pilih akun secara manual."),
            ("3️⃣", "Jalankan Pipeline",
             "Buka halaman Jalankan → klik 'Jalankan Semua'. Hasil tersimpan otomatis di folder ExtractedData_XBRL."),
        ]
        for ico, title, detail in steps_guide:
            row = ctk.CTkFrame(guide_card, fg_color=C_CARD2, corner_radius=8)
            row.pack(fill="x", pady=4)
            ri = ctk.CTkFrame(row, fg_color="transparent")
            ri.pack(fill="x", padx=14, pady=10)
            ctk.CTkLabel(ri, text=ico, font=ctk.CTkFont("Segoe UI", 18)).pack(side="left", padx=(0, 12))
            txt = ctk.CTkFrame(ri, fg_color="transparent")
            txt.pack(side="left", fill="x", expand=True)
            ctk.CTkLabel(txt, text=title,
                         font=ctk.CTkFont("Segoe UI", 11, "bold"),
                         text_color=C_WHITE, anchor="w").pack(anchor="w")
            ctk.CTkLabel(txt, text=detail,
                         font=ctk.CTkFont("Segoe UI", 9), text_color=C_MUTED2,
                         anchor="w", justify="left").pack(anchor="w", pady=(2, 0))

        # Spesifikasi teknis
        self._section_header(scroll, "⚙️", "Spesifikasi Teknis",
                              "Detail pipeline dan kemampuan aplikasi.")

        spec_card = self._card(scroll)
        specs = [
            ("🔗", "Sumber Data",    "idx.co.id — Inline XBRL Financial Statements (LK Tahunan & Kuartalan)"),
            ("📊", "Akun XBRL",      f"{len(ALL_METRICS)} akun tersedia dari 4 kategori: Informasi Umum, Neraca, Laba Rugi, Arus Kas"),
            ("🔄", "Pipeline",       "8 langkah: Generate → Download → Unzip → Ekstrak → Clean → Normalisasi → Export → Metrik"),
            ("⚗️", "Analisis",       "Altman Z-Score (prediksi kebangkrutan) · IBD — Interest Bearing Debt"),
            ("💾", "Output",         "2 file Excel: data mentah (RAW) dan data yang sudah dikonversi (Full IDR)"),
            ("🤖", "Anti-bot",       "Persistent session + header matching + retry otomatis untuk bypass proteksi IDX"),
        ]
        for ico, lbl, val in specs:
            r = ctk.CTkFrame(spec_card, fg_color="transparent")
            r.pack(fill="x", pady=5)
            ctk.CTkLabel(r, text=f"{ico}  {lbl}",
                         font=ctk.CTkFont("Segoe UI", 10, "bold"),
                         text_color=C_ACCENT, width=140, anchor="w").pack(side="left")
            ctk.CTkLabel(r, text=val,
                         font=ctk.CTkFont("Segoe UI", 10), text_color=C_TEXT,
                         wraplength=700, justify="left").pack(side="left")

        # Footer
        ctk.CTkFrame(scroll, fg_color=C_BORDER, height=1).pack(fill="x", pady=16)
        ctk.CTkLabel(scroll, text="IDX Superapp · Data Keuangan XBRL · idx.co.id",
                     font=ctk.CTkFont("Segoe UI", 9), text_color=C_MUTED).pack()

    # ═════════════════════════════════════════════════════════════════
    #  HELPER WIDGETS
    # ═════════════════════════════════════════════════════════════════
    def _page_header(self, parent, title: str, subtitle: str):
        hdr = ctk.CTkFrame(parent, fg_color=C_SURFACE, corner_radius=0)
        hdr.pack(fill="x")
        # Border bawah tipis
        ctk.CTkFrame(hdr, fg_color=C_BORDER, height=1,
                     corner_radius=0).pack(side="bottom", fill="x")
        hi = ctk.CTkFrame(hdr, fg_color="transparent")
        hi.pack(fill="x", padx=28, pady=(20, 18))
        ctk.CTkLabel(hi, text=title,
                     font=ctk.CTkFont("Segoe UI", 18, "bold"),
                     text_color=C_TEXT).pack(anchor="w")
        ctk.CTkLabel(hi, text=subtitle,
                     font=ctk.CTkFont("Segoe UI", 10),
                     text_color=C_MUTED).pack(anchor="w", pady=(2, 0))

    def _section_header(self, parent, badge: str, title: str, subtitle: str = ""):
        row = ctk.CTkFrame(parent, fg_color="transparent")
        row.pack(fill="x", pady=(20, 6))
        # Badge dengan bg accent tipis
        ctk.CTkLabel(row, text=badge,
                     font=ctk.CTkFont("Segoe UI", 10, "bold"),
                     text_color=C_ACCENT_LT,
                     fg_color="#1e1b4b", corner_radius=8,
                     width=26, height=26).pack(side="left", padx=(0, 10))
        txt = ctk.CTkFrame(row, fg_color="transparent")
        txt.pack(side="left")
        ctk.CTkLabel(txt, text=title,
                     font=ctk.CTkFont("Segoe UI", 12, "bold"),
                     text_color=C_TEXT).pack(anchor="w")
        if subtitle:
            ctk.CTkLabel(txt, text=subtitle,
                         font=ctk.CTkFont("Segoe UI", 9),
                         text_color=C_MUTED).pack(anchor="w")

    def _hint(self, parent, text: str):
        f = ctk.CTkFrame(parent, fg_color="#1c1917", corner_radius=6)
        f.pack(fill="x", pady=(8, 0))
        ctk.CTkLabel(f, text=f"  💡  {text}",
                     font=ctk.CTkFont("Segoe UI", 9),
                     text_color=C_MUTED2,
                     justify="left").pack(anchor="w", padx=4, pady=6)

    def _card(self, parent):
        frame = ctk.CTkFrame(parent, fg_color=C_CARD, corner_radius=10,
                              border_color=C_BORDER, border_width=1)
        frame.pack(fill="x", pady=(4, 10))
        inner = ctk.CTkFrame(frame, fg_color="transparent")
        inner.pack(fill="x", padx=20, pady=16)
        return inner

    # ═════════════════════════════════════════════════════════════════
    #  LOG & PROGRESS
    # ═════════════════════════════════════════════════════════════════
    def log(self, msg: str, tag: str = None):
        def _do():
            self.log_box.configure(state="normal")
            ts = datetime.now().strftime("%H:%M:%S")
            tb = self.log_box._textbox
            tb.insert("end", f"[{ts}] {msg}\n", tag or "")
            tb.see("end")
            self.log_box.configure(state="disabled")
        self.after(0, _do)

    def _clear_log(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("0.0", "end")
        self.log_box.configure(state="disabled")

    def _pb_set(self, canvas: tk.Canvas, value: float, color: str = C_ACCENT):
        """Update tk.Canvas progress bar tanpa re-layout. value: 0.0–1.0"""
        w = canvas.winfo_width() or 130
        fill_w = max(0, min(w, int(value * w)))
        canvas.coords("fill", 0, 0, fill_w, 6)
        canvas.itemconfig("fill", fill=color)

    def _set_step(self, key, status, color=None, pb_val=None):
        def _do():
            w = self._step_w.get(key)
            if not w: return
            c = color or C_MUTED
            # tk.Label native — update langsung, tanpa CTk re-draw
            w["lbl"].config(text=status, fg=c)
            if pb_val is not None:
                pb_color = C_SUCCESS if color == C_SUCCESS else (C_DANGER if color == C_DANGER else C_ACCENT)
                self._pb_set(w["pb"], pb_val, pb_color)
            # Badge nomor — CTkLabel kecil, update sekali saja
            if color == C_SUCCESS:
                w["num"].configure(fg_color=C_SUCCESS)
            elif color == C_DANGER:
                w["num"].configure(fg_color=C_DANGER)
            elif color == C_WARNING:
                w["num"].configure(fg_color=C_WARNING)
            else:
                w["num"].configure(fg_color=C_ACCENT)
        self.after(0, _do)

    def _set_progress(self, key, cur, tot):
        def _do():
            w = self._step_w.get(key)
            if w and tot > 0:
                self._pb_set(w["pb"], cur / tot, C_ACCENT)
        self.after(0, _do)

    def _set_overall(self, cur, tot, label=""):
        def _do():
            v = cur / tot if tot > 0 else 0
            pct = int(v * 100)
            # tk.Canvas overall progress — coords langsung, nol re-layout
            w = self._overall_pb_w or self.overall_pb.winfo_width() or 400
            fill_w = max(0, int(v * w))
            self.overall_pb.coords("fill", 0, 0, fill_w, 8)
            # tk.Label native
            self.lbl_overall_pct.config(text=f"{pct}%")
            if label: self.lbl_overall.config(text=label)
        self.after(0, _do)

    def _set_btns(self, running: bool):
        def _do():
            s = "disabled" if running else "normal"
            self.btn_run_all.configure(state=s)
            self.btn_stop.configure(
                state="normal" if running else "disabled",
                fg_color=C_DANGER if running else C_CARD2,
                text_color=C_WHITE if running else C_MUTED2,
            )
            # Tombol step adalah tk.Button (bukan CTkButton)
            for w in self._step_w.values():
                w["btn"].config(state=s)
        self.after(0, _do)

    def _stop_pipeline(self):
        self._stop_flag = True
        self.log("⏹  Stop diminta — menunggu langkah saat ini selesai...", "warn")

    # ═════════════════════════════════════════════════════════════════
    #  PIPELINE LOGIC
    # ═════════════════════════════════════════════════════════════════
    def _parse_usd_rate(self) -> float:
        raw = self.var_usd_rate.get().strip().replace(" ", "")
        if not raw:
            raise ValueError("Kurs USD kosong. Isi kurs USD ke IDR di halaman Pengaturan.")
        s = raw
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".") if s.rfind(",") > s.rfind(".") else s.replace(",", "")
        elif "," in s:
            s = s.replace(",", ".") if s.count(",") == 1 and len(s.split(",")[1]) <= 2 else s.replace(",", "")
        elif "." in s:
            if s.count(".") > 1:
                s = s.replace(".", "")
            else:
                left, right = s.split(".")
                if left.isdigit() and right.isdigit() and len(right) == 3:
                    s = left + right
        rate = float(s)
        if rate <= 0:
            raise ValueError("Kurs USD harus lebih besar dari 0.")
        return rate

    def _run_all(self):
        if self._running: return
        selected = self.get_selected_metrics()
        if not selected:
            messagebox.showwarning(
                "Akun Belum Dipilih",
                "Anda belum memilih akun yang akan diekstrak.\n\n"
                "Buka halaman 'Pilih Akun' dan pilih akun yang diinginkan,\n"
                "atau gunakan tombol preset Altman Z-Score / IBD."
            )
            return
        self._stop_flag = False
        self._running = True
        self._set_btns(True)
        self.log("=" * 60, "dim")
        self.log(f"▶  Pipeline dimulai — {len(selected)} akun dipilih", "info")

        def _go():
            step_names = {
                "generate": "Generate Links", "download": "Download XBRL",
                "unzip": "Ekstrak ZIP",        "extract": "Ekstrak XBRL",
                "normalize": "Normalisasi IDR",
                "export": "Export Excel",      "metrics": "Hitung Metrik",
            }
            keys = list(step_names.keys())
            n = len(keys)
            completed = 0
            for i, key in enumerate(keys):
                if self._stop_flag: break
                self._set_overall(i, n, f"Langkah {i+1}/{n} — {step_names[key]}")
                self._do_step(key)
                completed = i + 1

            if self._stop_flag:
                self._set_overall(completed, n, f"⏹  Dihentikan setelah {completed}/{n} langkah")
            else:
                self._set_overall(n, n, "✅  Semua langkah selesai!")

            self._running = False
            self._set_btns(False)
            if not self._stop_flag:
                self.log("✅  Pipeline selesai! Cek folder ExtractedData_XBRL untuk hasilnya.", "ok")
                self.after(0, lambda: messagebox.showinfo(
                    "Pipeline Selesai! ✅",
                    "Ekstraksi data berhasil diselesaikan.\n\n"
                    "📁 Hasil tersimpan di folder ExtractedData_XBRL/:\n"
                    "   • 00_SUMMARY_all_companies_raw.xlsx  (data mentah)\n"
                    "   • 00_SUMMARY_all_companies_full_idr.xlsx  (nilai IDR penuh)\n"
                    "   • 00_SUMMARY_all_companies_full_idr_metrics.xlsx  (+ Altman/IBD)"
                ))
        threading.Thread(target=_go, daemon=True).start()

    def _run_step(self, key: str):
        if self._running: return
        self._stop_flag = False
        self._running = True
        self._set_btns(True)
        all_keys = ["generate", "download", "unzip", "extract", "normalize", "export", "metrics"]
        n = len(all_keys)
        idx = all_keys.index(key) if key in all_keys else 0
        step_names = {
            "generate": "Generate Links", "download": "Download XBRL",
            "unzip": "Ekstrak ZIP",        "extract": "Ekstrak XBRL",
            "normalize": "Normalisasi IDR",
            "export": "Export Excel",      "metrics": "Hitung Metrik",
        }
        self._set_overall(idx, n, f"Langkah {idx+1}/{n} — {step_names.get(key, key)}")
        def _go():
            self._do_step(key)
            # Update progress ke selesai step ini
            self._set_overall(idx + 1, n, f"✓ Langkah {idx+1}/{n} selesai — {step_names.get(key, key)}")
            self._running = False
            self._set_btns(False)
        threading.Thread(target=_go, daemon=True).start()

    def _do_step(self, key: str):
        year  = self.var_year.get()
        period = self.var_period.get()
        cf    = self.var_companies_file.get()
        dl    = self.var_download_dir.get()
        xbrl  = self.var_xbrl_dir.get()
        ext   = self.var_extract_dir.get()
        csv_s        = os.path.join(ext, "00_SUMMARY_all_companies.csv")
        csv_raw      = os.path.join(ext, "00_SUMMARY_all_companies_raw.csv")
        csv_full_idr = os.path.join(ext, "00_SUMMARY_all_companies_full_idr.csv")

        slog = lambda m: self.log(m)
        spg  = lambda c, t: self._set_progress(key, c, t)
        stop = lambda: self._stop_flag

        self._set_step(key, "Berjalan...", C_WARNING)
        try:
            if key == "generate":
                self.log(f"🔗 Generate links — Tahun {year}, {period}", "info")
                n = generate_links(year, period, cf, progress_callback=spg, log_callback=slog)
                if n and n > 0:
                    self._set_step(key, f"✓ {n} link", C_SUCCESS, 1.0)
                    self.log(f"✓ {n} link disimpan → {cf}", "ok")
                else:
                    self._set_step(key, "Gagal", C_DANGER)

            elif key == "download":
                self.log("⬇️  Mengunduh file XBRL dari IDX...", "info")
                st = download_all(cf, dl, spg, slog, stop)
                self._set_step(key, f"✓ {st['success']} file", C_SUCCESS, 1.0)
                self.log(f"✓ {st['success']} berhasil · {st['skipped']} dilewati · {st['bot_detected']} bot", "ok")

            elif key == "unzip":
                self.log("📦  Mengekstrak file ZIP...", "info")
                st = unzip_all(dl, xbrl, spg, slog, stop)
                self._set_step(key, f"✓ {st['success']} file", C_SUCCESS, 1.0)

            elif key == "extract":
                selected = self.get_selected_metrics()
                if not selected:
                    self.log("⚠  Tidak ada akun dipilih! Buka halaman Pilih Akun.", "warn")
                    self._set_step(key, "Tidak ada akun", C_WARNING)
                    return
                selected_for_extract = list(selected)
                added_refs = []
                for ref in NORMALIZATION_REFERENCE_METRICS:
                    if ref not in selected_for_extract:
                        selected_for_extract.append(ref)
                        added_refs.append(ref)
                if added_refs:
                    self.log("ℹ  Menambahkan akun referensi normalisasi: " + ", ".join(added_refs), "dim")
                self.log(f"🔍  Mengekstrak {len(selected_for_extract)} akun dari file XBRL...", "info")
                r = extract_all(xbrl, ext, selected_metrics=selected_for_extract,
                                progress_callback=spg, log_callback=slog, stop_flag=stop)
                if r:
                    self._set_step(key, "✓ CSV dibuat", C_SUCCESS, 1.0)
                    self.log(f"✓ CSV → {r}", "ok")
                else:
                    self._set_step(key, "Gagal/kosong", C_WARNING)

            elif key == "clean":
                self.log("🧹  Membersihkan nilai negatif pada akun beban bunga...", "info")
                ok = clean_data(csv_s, log_callback=slog)
                self._set_step(key, "✓ Selesai" if ok else "Gagal",
                               C_SUCCESS if ok else C_DANGER, 1.0)

            elif key == "normalize":
                selected = self.get_selected_metrics()
                usd_rate = self._parse_usd_rate()
                self.log(f"💱  Normalisasi ke Rupiah penuh (kurs USD = {usd_rate:,.0f})...", "info")
                if not os.path.exists(csv_s):
                    raise FileNotFoundError(f"File CSV tidak ditemukan: {csv_s}\nJalankan langkah Ekstrak XBRL terlebih dahulu.")
                try:
                    shutil.copy2(csv_s, csv_raw)
                    self.log(f"✓ Snapshot RAW disimpan → {csv_raw}", "dim")
                except Exception as e:
                    raise RuntimeError(f"Gagal membuat snapshot RAW: {e}")
                ok = normalize_to_full_idr(csv_s, selected_metrics=selected,
                                           usd_rate=usd_rate, log_callback=slog)
                if ok:
                    try:
                        shutil.copy2(csv_s, csv_full_idr)
                        self.log(f"✓ Hasil Full IDR disimpan → {csv_full_idr}", "dim")
                    except Exception as e:
                        raise RuntimeError(f"Normalisasi berhasil, tapi gagal simpan: {e}")
                self._set_step(key, "✓ Selesai" if ok else "Gagal",
                               C_SUCCESS if ok else C_DANGER, 1.0)

            elif key == "export":
                self.log("📊  Mengekspor ke Excel...", "info")
                if not os.path.exists(csv_s):
                    raise FileNotFoundError(f"CSV tidak ditemukan: {csv_s}")
                if not os.path.exists(csv_raw):
                    shutil.copy2(csv_s, csv_raw)
                    self.log("ℹ  File RAW belum ada, dibuat dari CSV terkini.", "warn")
                if not os.path.exists(csv_full_idr):
                    shutil.copy2(csv_s, csv_full_idr)
                    self.log("ℹ  File Full IDR belum ada, dibuat dari CSV terkini.", "warn")
                raw_xlsx     = os.path.join(ext, "00_SUMMARY_all_companies_raw.xlsx")
                full_idr_xlsx = os.path.join(ext, "00_SUMMARY_all_companies_full_idr.xlsx")
                xlsx_raw      = export_to_excel(csv_raw,      excel_path=raw_xlsx,      log_callback=slog)
                xlsx_full_idr = export_to_excel(csv_full_idr, excel_path=full_idr_xlsx, log_callback=slog)
                if xlsx_raw and xlsx_full_idr:
                    self._set_step(key, "✓ 2 Excel", C_SUCCESS, 1.0)
                    self.log(f"✓ RAW Excel    → {xlsx_raw}", "ok")
                    self.log(f"✓ Full IDR     → {xlsx_full_idr}", "ok")
                else:
                    self._set_step(key, "Gagal", C_DANGER)

            elif key == "metrics":
                selected_set = set(self.get_selected_metrics())
                run_altman = bool(selected_set & set(self.ALTMAN_ACCOUNTS))
                run_ibd    = bool(selected_set & set(self.IBD_ACCOUNTS))
                if not run_altman and not run_ibd:
                    self.log("⚗️  Tidak ada akun Altman/IBD dipilih — langkah ini dilewati.\n"
                             "   Gunakan tombol preset di halaman Pilih Akun.", "warn")
                    self._set_step(key, "Dilewati", C_MUTED, 1.0)
                else:
                    full_idr_xlsx = os.path.join(ext, "00_SUMMARY_all_companies_full_idr.xlsx")
                    metrics_xlsx  = os.path.join(ext, "00_SUMMARY_all_companies_full_idr_metrics.xlsx")
                    what = []
                    if run_altman: what.append("Altman Z-Score")
                    if run_ibd:    what.append("IBD")
                    self.log(f"⚗️  Menghitung: {', '.join(what)} ...", "info")
                    if not os.path.exists(full_idr_xlsx):
                        raise FileNotFoundError(
                            f"File Excel Full IDR tidak ditemukan: {full_idr_xlsx}\n"
                            "Jalankan langkah Export terlebih dahulu.")
                    ok = calculate_metrics(
                        xlsx_path=full_idr_xlsx, output_path=metrics_xlsx,
                        run_altman=run_altman, run_ibd=run_ibd, log_callback=slog)
                    if ok:
                        self._set_step(key, "✓ Selesai", C_SUCCESS, 1.0)
                        self.log(f"✓ Metrics Excel → {metrics_xlsx}", "ok")
                    else:
                        self._set_step(key, "Gagal", C_DANGER)

        except Exception as e:
            self._set_step(key, "Error!", C_DANGER)
            self.log(f"✗  Error pada langkah '{key}': {e}", "err")


def run():
    app = IDXSuperApp()
    app.mainloop()


if __name__ == "__main__":
    run()
