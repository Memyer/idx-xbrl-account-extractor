"""
core/metric_calculator.py
Hitung metrik keuangan turunan dari data XBRL yang sudah diekstrak + dinormalisasi.

Fitur:
  1. Altman Z-Score (modified for non-financial firms)
     - Preprocessing kolom pendapatan & beban bunga
     - Penambahan Retained_Earnings
     - Rumus Excel di setiap kolom turunan (bukan nilai statis)
  2. Interest Bearing Debt (IBD)
     - Menjumlahkan semua komponen utang berbunga
     - Rumus Excel di kolom IBD
"""
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Optional


# ── Kolom sumber (nama asli dari XBRL) ────────────────────────────────────
COL_REVENUE       = "Penjualan dan pendapatan usaha"
COL_PENDAPATAN_BUNGA = "Pendapatan bunga"
COL_PENDAPATAN_OP  = "Pendapatan operasional lainnya"
COL_BEBAN_BUNGA_KAK = "Beban bunga dan keuangan"
COL_BEBAN_BUNGA    = "Beban bunga"
COL_SALDO_LABA_1   = "Saldo laba yang belum ditentukan penggunaannya"
COL_SALDO_LABA_2   = "Saldo laba yang telah ditentukan penggunaannya"
COL_EBT            = "Jumlah laba (rugi) sebelum pajak penghasilan"
COL_TOTAL_ASET     = "Jumlah aset"
COL_ASET_LANCAR    = "Jumlah aset lancar"
COL_LIB_PENDEK     = "Jumlah liabilitas jangka pendek"
COL_EKUITAS        = "Jumlah ekuitas"
COL_DEPRESIASI     = "Depresiasi"
COL_AMORTISASI     = "Amortisasi"

# ── Warna header untuk kolom turunan ──────────────────────────────────────
FILL_ALTMAN = PatternFill("solid", fgColor="1A3A5C")   # biru gelap
FILL_IBD    = PatternFill("solid", fgColor="1A4A2A")   # hijau gelap
FONT_HEADER = Font(bold=True, color="FFFFFF", size=9)
FONT_FORMULA = Font(color="E0E8FF", size=9)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center")


# ── Helpers ────────────────────────────────────────────────────────────────
def _col_letter(df: pd.DataFrame, col_name: str) -> Optional[str]:
    """Kembalikan huruf kolom Excel (A, B, …) untuk nama kolom DataFrame."""
    if col_name not in df.columns:
        return None
    idx = list(df.columns).index(col_name) + 1   # openpyxl 1-based
    return get_column_letter(idx)


def _append_col(df: pd.DataFrame, name: str, default=None) -> str:
    """Tambahkan kolom kosong ke df dan kembalikan letter-nya."""
    df[name] = default
    return get_column_letter(len(df.columns))


def _numeric(df: pd.DataFrame, col: str) -> bool:
    return col in df.columns and pd.api.types.is_numeric_dtype(df[col])


def calculate_metrics(
    xlsx_path: str,
    output_path: str,
    run_altman: bool = True,
    run_ibd: bool = True,
    log_callback=None,
) -> bool:
    """
    Baca xlsx_path, tambahkan kolom metrik dengan rumus Excel, simpan ke output_path.

    Preprocessing yang dilakukan sebelum kalkulasi:
    - Jika COL_REVENUE kosong untuk suatu baris → isi dari pendapatan_bunga + pendapatan_op
      lalu hapus kedua kolom sumber.
    - Jika COL_BEBAN_BUNGA_KAK kosong → isi dari COL_BEBAN_BUNGA, lalu hapus kolom sumber.
    - Tambahkan kolom Retained_Earnings = saldo_laba_1 + saldo_laba_2.

    Returns True jika berhasil, False jika gagal.
    """
    def log(msg):
        if log_callback: log_callback(msg)
        else: print(msg)

    if not os.path.exists(xlsx_path):
        log(f"✗ File tidak ditemukan: {xlsx_path}")
        return False

    try:
        log(f"  Membaca: {xlsx_path}")
        ext = os.path.splitext(xlsx_path)[1].lower()
        if ext == ".csv":
            df = pd.read_csv(xlsx_path, encoding="utf-8-sig")
        else:
            df = pd.read_excel(xlsx_path, sheet_name=0)
        log(f"  → {len(df)} baris, {len(df.columns)} kolom")
    except Exception as e:
        log(f"✗ Gagal membaca file: {e}")
        return False

    # ── Preprocessing ──────────────────────────────────────────────────────
    drop_cols = []

    # 1. Revenue: isi dari pendapatan_bunga + pendapatan_op jika kosong
    if COL_REVENUE in df.columns:
        if COL_PENDAPATAN_BUNGA in df.columns and COL_PENDAPATAN_OP in df.columns:
            mask_empty = df[COL_REVENUE].isna() | (df[COL_REVENUE] == 0)
            pb = pd.to_numeric(df[COL_PENDAPATAN_BUNGA], errors="coerce").fillna(0)
            po = pd.to_numeric(df[COL_PENDAPATAN_OP],    errors="coerce").fillna(0)
            df.loc[mask_empty, COL_REVENUE] = pb[mask_empty] + po[mask_empty]
            drop_cols += [COL_PENDAPATAN_BUNGA, COL_PENDAPATAN_OP]
            log(f"  Preprocessing: '{COL_REVENUE}' — isi dari pendapatan_bunga + pendapatan_op lalu hapus kolom sumber")

    # 2. Beban bunga & keuangan: isi dari beban_bunga jika kosong
    if COL_BEBAN_BUNGA_KAK in df.columns and COL_BEBAN_BUNGA in df.columns:
        mask_empty = df[COL_BEBAN_BUNGA_KAK].isna() | (df[COL_BEBAN_BUNGA_KAK] == 0)
        bb = pd.to_numeric(df[COL_BEBAN_BUNGA], errors="coerce").fillna(0)
        df.loc[mask_empty, COL_BEBAN_BUNGA_KAK] = bb[mask_empty]
        drop_cols.append(COL_BEBAN_BUNGA)
        log(f"  Preprocessing: '{COL_BEBAN_BUNGA_KAK}' — isi dari beban_bunga lalu hapus kolom sumber")

    # Hapus kolom sumber yang sudah di-merge
    actual_drop = [c for c in drop_cols if c in df.columns]
    if actual_drop:
        df.drop(columns=actual_drop, inplace=True)

    # 3. Retained_Earnings
    re_col = "Retained_Earnings"
    if COL_SALDO_LABA_1 in df.columns or COL_SALDO_LABA_2 in df.columns:
        sl1 = pd.to_numeric(df.get(COL_SALDO_LABA_1, 0), errors="coerce").fillna(0)
        sl2 = pd.to_numeric(df.get(COL_SALDO_LABA_2, 0), errors="coerce").fillna(0)
        df[re_col] = sl1 + sl2
        log(f"  Preprocessing: kolom '{re_col}' ditambahkan (saldo_laba_1 + saldo_laba_2)")

    # ── Simpan dulu ke file agar bisa ditulis ulang dengan openpyxl ───────
    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Data Keuangan")
        log(f"  Data dasar tersimpan ke: {output_path}")
    except Exception as e:
        log(f"✗ Gagal simpan Excel awal: {e}")
        return False

    # ── Buka kembali dengan openpyxl untuk menambah rumus ─────────────────
    try:
        wb = load_workbook(output_path)
        ws = wb.active
        nrows = ws.max_row          # termasuk header (row 1)
        data_rows = nrows - 1       # jumlah baris data

        # Refresh daftar kolom (mungkin sudah berubah karena preprocessing)
        header_row = [cell.value for cell in ws[1]]
        col_map = {name: get_column_letter(i + 1) for i, name in enumerate(header_row) if name}

        def L(name):
            """Kembalikan letter kolom untuk nama kolom, atau None jika tidak ada."""
            return col_map.get(name)

        # ── IBD ──────────────────────────────────────────────────────────
        if run_ibd:
            ibd_components = [
                # Utang jangka pendek berbunga
                'Utang bank jangka pendek',
                'Utang trust receipts',
                'Pinjaman jangka pendek non-bank',
                'Utang lembaga keuangan non-bank',
                # Jatuh tempo dalam 1 tahun
                'Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas utang bank',
                'Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas utang keuangan keuangan non bank',
                'Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas utang obligasi',
                'Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas surat utang jangka menengah',
                'Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas utang pembiayaan konsumen',
                'Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas liabilitas sewa pembiayaan',
                'Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas sukuk',
                'Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas obligasi subordinasi',
                'Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas pinjaman beragunan',
                'Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas pinjaman tanpa agunan',
                'Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas penerusan pinjaman',
                'Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas pinjaman dari pemerintah republik Indonesia',
                'Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas pinjaman subordinasi',
                'Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas liabilitas kerja sama operasi',
                'Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas liabilitas pembebasan tanah',
                'Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas utang listrik swasta',
                'Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas utang retensi',
                'Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas wesel bayar',
                'Liabilitas jangka panjang yang jatuh tempo dalam satu tahun atas pinjaman lainnya',
                # Utang jangka panjang
                'Liabilitas jangka panjang atas utang bank',
                'Liabilitas jangka panjang atas utang obligasi',
                'Liabilitas jangka panjang atas pinjaman beragunan',
                'Liabilitas jangka panjang atas pinjaman tanpa agunan',
                'Liabilitas jangka panjang atas pinjaman dari pemerintah republik Indonesia',
                'Liabilitas jangka panjang atas pinjaman subordinasi',
                'Liabilitas jangka panjang atas utang pembiayaan konsumen',
                'Liabilitas jangka panjang atas surat utang jangka menengah',
                'Liabilitas jangka panjang atas sukuk',
                'Liabilitas jangka panjang atas obligasi subordinasi',
                'Liabilitas jangka panjang atas liabilitas sewa pembiayaan',
                'Liabilitas jangka panjang atas penerusan pinjaman',
                'Liabilitas jangka panjang atas liabilitas kerja sama operasi',
                'Liabilitas jangka panjang atas liabilitas pembebasan tanah',
                'Liabilitas jangka panjang atas utang listrik swasta',
                'Liabilitas jangka panjang atas utang retensi',
                'Liabilitas jangka panjang atas wesel bayar',
                'Liabilitas jangka panjang atas pinjaman lainnya',
                # Efek utang yang diterbitkan (neraca)
                'Utang obligasi',
                'Sukuk',
                'Obligasi subordinasi',
                'Liabilitas sewa pembiayaan',
                'Obligasi konversi',
                'Pinjaman subordinasi pihak ketiga',
                'Pinjaman subordinasi pihak berelasi',
                'Sukuk mudharabah',
                'Sukuk mudharabah subordinasi',
            ]
            present = [c for c in ibd_components if L(c)]
            if present:
                # Tambah header kolom IBD di akhir
                ibd_col_idx = ws.max_column + 1
                ibd_col_ltr = get_column_letter(ibd_col_idx)
                hdr_cell = ws.cell(row=1, column=ibd_col_idx, value="IBD")
                hdr_cell.fill = FILL_IBD
                hdr_cell.font = FONT_HEADER
                hdr_cell.alignment = ALIGN_CENTER

                for r in range(2, nrows + 1):
                    sum_parts = "+".join(f"IFERROR({L(c)}{r}*1,0)" for c in present)
                    formula = f"={sum_parts}"
                    cell = ws.cell(row=r, column=ibd_col_idx, value=formula)
                    cell.font = FONT_FORMULA
                    cell.alignment = ALIGN_RIGHT
                log(f"  ✓ Kolom IBD ditambahkan ({len(present)} komponen)")
            else:
                log(f"  ⚠ IBD: tidak ada kolom komponen yang ditemukan di data")

        # ── Altman Z-Score ────────────────────────────────────────────────
        if run_altman:
            # Kolom yang dibutuhkan
            ebt_l     = L(COL_EBT)
            bbk_l     = L(COL_BEBAN_BUNGA_KAK)
            dep_l     = L(COL_DEPRESIASI)
            amor_l    = L(COL_AMORTISASI)
            ta_l      = L(COL_TOTAL_ASET)
            rev_l     = L(COL_REVENUE)
            eq_l      = L(COL_EKUITAS)
            al_l      = L(COL_ASET_LANCAR)
            lp_l      = L(COL_LIB_PENDEK)
            re_l      = L(re_col)

            next_col = ws.max_column + 1

            def add_header(col_idx, title, fill=FILL_ALTMAN):
                c = ws.cell(row=1, column=col_idx, value=title)
                c.fill = fill
                c.font = FONT_HEADER
                c.alignment = ALIGN_CENTER
                return get_column_letter(col_idx)

            def safe(letter):
                """Wrap letter reference untuk menghindari div/0 dan error."""
                return f"IFERROR({letter}*1,0)" if letter else "0"

            # --- EBIT ---
            ebit_idx = next_col
            ebit_ltr = add_header(ebit_idx, "EBIT")
            for r in range(2, nrows + 1):
                if ebt_l and bbk_l:
                    f = f"=IFERROR({ebt_l}{r},0)+IFERROR({bbk_l}{r},0)"
                elif ebt_l:
                    f = f"=IFERROR({ebt_l}{r},0)"
                else:
                    f = "="
                cell = ws.cell(row=r, column=ebit_idx, value=f)
                cell.font = FONT_FORMULA; cell.alignment = ALIGN_RIGHT
            next_col += 1

            # --- EBITDA ---
            ebitda_idx = next_col
            ebitda_ltr = add_header(ebitda_idx, "EBITDA")
            for r in range(2, nrows + 1):
                dep_part  = f"+IFERROR({dep_l}{r},0)"  if dep_l  else ""
                amor_part = f"+IFERROR({amor_l}{r},0)" if amor_l else ""
                f = f"={ebit_ltr}{r}{dep_part}{amor_part}"
                cell = ws.cell(row=r, column=ebitda_idx, value=f)
                cell.font = FONT_FORMULA; cell.alignment = ALIGN_RIGHT
            next_col += 1

            # --- EBIT_TA ---
            ebit_ta_idx = next_col
            ebit_ta_ltr = add_header(ebit_ta_idx, "EBIT_TA")
            for r in range(2, nrows + 1):
                if ta_l:
                    f = f"=IFERROR({ebit_ltr}{r}/IF({ta_l}{r}=0,NA(),{ta_l}{r}),\"\")"
                else:
                    f = "=\"\""
                cell = ws.cell(row=r, column=ebit_ta_idx, value=f)
                cell.font = FONT_FORMULA; cell.alignment = ALIGN_RIGHT
            next_col += 1

            # --- Revenue_TA ---
            rev_ta_idx = next_col
            rev_ta_ltr = add_header(rev_ta_idx, "Revenue_TA")
            for r in range(2, nrows + 1):
                if rev_l and ta_l:
                    f = f"=IFERROR(IFERROR({rev_l}{r},0)/IF({ta_l}{r}=0,NA(),{ta_l}{r}),\"\")"
                else:
                    f = "=\"\""
                cell = ws.cell(row=r, column=rev_ta_idx, value=f)
                cell.font = FONT_FORMULA; cell.alignment = ALIGN_RIGHT
            next_col += 1

            # --- Equity_TA_minus_Equity ---
            eq_ratio_idx = next_col
            eq_ratio_ltr = add_header(eq_ratio_idx, "Equity_TA_minus_Equity")
            for r in range(2, nrows + 1):
                if eq_l and ta_l:
                    f = (f"=IFERROR(IFERROR({eq_l}{r},0)/"
                         f"IF(({ta_l}{r}-IFERROR({eq_l}{r},0))=0,NA(),"
                         f"({ta_l}{r}-IFERROR({eq_l}{r},0))),\"\")")
                else:
                    f = "=\"\""
                cell = ws.cell(row=r, column=eq_ratio_idx, value=f)
                cell.font = FONT_FORMULA; cell.alignment = ALIGN_RIGHT
            next_col += 1

            # --- WC_TA ---
            wc_ta_idx = next_col
            wc_ta_ltr = add_header(wc_ta_idx, "WC_TA")
            for r in range(2, nrows + 1):
                if al_l and lp_l and ta_l:
                    f = (f"=IFERROR((IFERROR({al_l}{r},0)-IFERROR({lp_l}{r},0))/"
                         f"IF({ta_l}{r}=0,NA(),{ta_l}{r}),\"\")")
                else:
                    f = "=\"\""
                cell = ws.cell(row=r, column=wc_ta_idx, value=f)
                cell.font = FONT_FORMULA; cell.alignment = ALIGN_RIGHT
            next_col += 1

            # --- RE_TA ---
            re_ta_idx = next_col
            re_ta_ltr = add_header(re_ta_idx, "RE_TA")
            for r in range(2, nrows + 1):
                if re_l and ta_l:
                    f = f"=IFERROR(IFERROR({re_l}{r},0)/IF({ta_l}{r}=0,NA(),{ta_l}{r}),\"\")"
                else:
                    f = "=\"\""
                cell = ws.cell(row=r, column=re_ta_idx, value=f)
                cell.font = FONT_FORMULA; cell.alignment = ALIGN_RIGHT
            next_col += 1

            # --- Altman Z-Score ---
            z_idx = next_col
            add_header(z_idx, "Altman_Z_Score")
            for r in range(2, nrows + 1):
                f = (f"=IFERROR("
                     f"3.107*IFERROR({ebit_ta_ltr}{r},0)"
                     f"+0.998*IFERROR({rev_ta_ltr}{r},0)"
                     f"+0.42*IFERROR({eq_ratio_ltr}{r},0)"
                     f"+0.717*IFERROR({wc_ta_ltr}{r},0)"
                     f"+0.847*IFERROR({re_ta_ltr}{r},0)"
                     f",\"\")")
                cell = ws.cell(row=r, column=z_idx, value=f)
                cell.font = FONT_FORMULA; cell.alignment = ALIGN_RIGHT
            log(f"  ✓ Kolom Altman Z-Score ditambahkan (EBIT, EBITDA, EBIT_TA, Revenue_TA, Equity_TA_minus_Equity, WC_TA, RE_TA, Altman_Z_Score)")

        # ── Auto-fit kolom ─────────────────────────────────────────────────
        for col_cells in ws.columns:
            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0)
                for cell in col_cells
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 40)

        wb.save(output_path)
        log(f"  ✓ Disimpan: {output_path}")
        return True

    except Exception as e:
        log(f"✗ Error saat kalkulasi metrik: {e}")
        import traceback
        log(traceback.format_exc())
        return False
